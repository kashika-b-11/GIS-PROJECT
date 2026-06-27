from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


CATEGORY_BY_DEPARTMENT = {
    "AYUSH": "health",
    "DIRECTORATE OF HEALTH": "health",
    "DIRECTORATE OF WOMEN & CHILD": "health",
    "HP TAKNIKI SHIKSHA BOARD": "education",
    "TECHNICAL EDUCATION,VOCATIONAL": "education",
    "POLICE DEPARTMENT": "safety",
    "STATE VIGILANCE BUREAU": "safety",
    "DIRECTORATE OF PRISONS": "safety",
    "FOREST CORPORATION": "forest",
    "FISHERIES": "livelihood",
    "PUBLIC WORKS": "infrastructure",
    "PANCHAYATI RAJ": "infrastructure",
    "RURAL DEVELOPMENT": "infrastructure",
    "DD&TG": "infrastructure",
    "State Excise Commissioner": "governance",
    "ELECTION": "governance",
    "ENVIRONMENT, SCIENCE": "governance",
}


def value_or_none(value):
    if value in (None, "", "NULL"):
        return None
    return value


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python tools/import_asset_workbook.py <input.xlsx> <output.geojson>")
        return 1

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    features = []

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]

        for row in rows:
            record = dict(zip(headers, row))
            lat = value_or_none(record.get("Latitude"))
            lng = value_or_none(record.get("Longitude"))
            if lat is None or lng is None:
                continue

            department = value_or_none(record.get("department_name")) or sheet_name
            feature = {
                "type": "Feature",
                "geometry": {
                    "type": "Point",
                    "coordinates": [float(lng), float(lat)],
                },
                "properties": {
                    "asset_name": value_or_none(record.get("asset_name")) or "Unnamed asset",
                    "asset_type_name": value_or_none(record.get("asset_type_name")),
                    "asset_area": value_or_none(record.get("asset_area")),
                    "asset_address": value_or_none(record.get("asset_address")),
                    "asset_description": value_or_none(record.get("asset_description")),
                    "department_name": department,
                    "department_sheet": sheet_name,
                    "asset_category": CATEGORY_BY_DEPARTMENT.get(sheet_name, "other"),
                    "district_name": value_or_none(record.get("district_lgd_name")),
                    "block_name": value_or_none(record.get("block_lgd_name")),
                    "panchayat_name": value_or_none(record.get("panchayat_lgd_name")),
                    "village_name": value_or_none(record.get("village_lgd_name")),
                    "office_name": value_or_none(record.get("office_name")),
                    "latitude": float(lat),
                    "longitude": float(lng),
                },
            }
            features.append(feature)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps({"type": "FeatureCollection", "features": features}, ensure_ascii=True),
        encoding="utf-8",
    )

    counts = {}
    for feature in features:
        category = feature["properties"]["asset_category"]
        counts[category] = counts.get(category, 0) + 1

    print(f"Wrote {len(features)} assets to {output_path}")
    for category, count in sorted(counts.items()):
        print(f"{category}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
